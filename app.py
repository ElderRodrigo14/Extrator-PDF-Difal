import streamlit as st
import pdfplumber
import openpyxl
import re
import io

st.set_page_config(page_title="Extrator DIFAL", layout="centered")

st.title("📄 Extrator de DIFAL/FCP para Excel")
st.write("Faça o upload de uma ou mais notas fiscais (PDF). O sistema extrairá os valores e gerará o relatório Excel preenchido.")

# Permite subir vários PDFs de uma vez
arquivos_pdf = st.file_uploader("Selecione os PDFs das Notas Fiscais", type=["pdf"], accept_multiple_files=True)

def extrair_dados_pdf(arquivo):
    texto_completo = ""
    with pdfplumber.open(arquivo) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if texto:
                texto_completo += texto + "\n"
    
    # 1. Buscar Número da NF (Ex: "N. 2.428")
    match_nf = re.search(r'N\.\s*([\d\.]+)', texto_completo)
    numero_nf = match_nf.group(1) if match_nf else "N/A"
    
    # 2. Buscar valor do DIFAL (Ex: "DIFAL da UF Destino R$21,08")
    match_difal = re.search(r'DIFAL da UF Destino R\$\s*([\d\.,]+)', texto_completo)
    valor_difal = 0.0
    if match_difal:
        # Troca vírgula por ponto para o Python entender como decimal
        valor_str = match_difal.group(1).replace('.', '').replace(',', '.')
        valor_difal = float(valor_str)
        
    # 3. Buscar valor do FCP (Ex: "FCP R$0,00")
    match_fcp = re.search(r'FCP R\$\s*([\d\.,]+)', texto_completo)
    valor_fcp = 0.0
    if match_fcp:
        valor_str = match_fcp.group(1).replace('.', '').replace(',', '.')
        valor_fcp = float(valor_str)
        
    valor_total = valor_difal + valor_fcp
    
    return numero_nf, valor_total

def preencher_excel(dados, caminho_modelo):
    # Carrega a planilha modelo
    wb = openpyxl.load_workbook(caminho_modelo)
    planilha = wb.active
    
    col_desc = None
    col_val = None
    linha_inicio = None
    
    # Procura em qual coluna estão "DESCRIÇÃO" e "VALOR" automaticamente
    for r in range(1, 20):
        for c in range(1, 10):
            celula = planilha.cell(row=r, column=c).value
            if celula == "DESCRIÇÃO":
                col_desc = c
                linha_inicio = r + 1
            elif celula == "VALOR":
                col_val = c
                
    if not col_desc or not col_val:
        st.error("Não foi possível encontrar os cabeçalhos 'DESCRIÇÃO' e 'VALOR' no Excel.")
        return None

    # Preenche os dados
    linha_atual = linha_inicio
    for nf, valor in dados:
        # Desce as linhas até achar um campo escrito "DIFAL" ou vazio
        val_atual = str(planilha.cell(row=linha_atual, column=col_desc).value).strip()
        while val_atual not in ["DIFAL", "None", ""]:
            linha_atual += 1
            val_atual = str(planilha.cell(row=linha_atual, column=col_desc).value).strip()
            
        # Insere a descrição formatada e o valor
        planilha.cell(row=linha_atual, column=col_desc).value = f"DIFAL NF {nf}"
        planilha.cell(row=linha_atual, column=col_val).value = valor
        linha_atual += 1

    # Salva na memória para permitir o download no site
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if arquivos_pdf:
    st.info(f"Processando {len(arquivos_pdf)} arquivo(s)...")
    
    dados_extraidos = []
    for arquivo in arquivos_pdf:
        nf, valor_total = extrair_dados_pdf(arquivo)
        if nf != "N/A":
            dados_extraidos.append((nf, valor_total))
            st.write(f"✅ **NF {nf}:** R$ {valor_total:.2f} (DIFAL + FCP)")
        else:
            st.warning(f"⚠️ Não foi possível encontrar o número da NF no arquivo: {arquivo.name}")
            
    if dados_extraidos:
        # Nome EXATO do seu arquivo base, conforme seu print
        nome_arquivo_excel = "Extrator PDF DIFAL.xlsx" 
        
        try:
            excel_preenchido = preencher_excel(dados_extraidos, nome_arquivo_excel)
            
            if excel_preenchido:
                st.success("Relatório gerado com sucesso!")
                
                # Botão para baixar o arquivo pronto
                st.download_button(
                    label="📥 Baixar Relatório Preenchido",
                    data=excel_preenchido,
                    file_name="Protocolo_DIFAL_Preenchido.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except FileNotFoundError:
            st.error(f"Erro: O arquivo '{nome_arquivo_excel}' não foi encontrado na mesma pasta do app.py.")
